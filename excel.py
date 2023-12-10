import openpyxl

workbook = openpyxl.load_workbook("P2-001_エンティティ定義書_マスタ.xlsx")
print(workbook.sheetnames)

txtFile = open(file="sheet_list.txt", mode="w")

for name in workbook.sheetnames:
	txtFile.write(name + "\r\n")

txtFile.close()
