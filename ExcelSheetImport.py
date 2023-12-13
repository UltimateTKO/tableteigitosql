"""
テーブル定義からシート一覧ファイルを出力
"""

import openpyxl
import os

xlsxFiles = [file for file in os.listdir(".") if ".xlsx" in os.path.join(".", file)]
workbook = openpyxl.load_workbook(xlsxFiles[0])
print(workbook.sheetnames)

txtFile = open(file="sheet_list.txt", mode="w")

for name in workbook.sheetnames:
	txtFile.write(name + "\n")

txtFile.close()
